# -*- coding: utf-8 -*-
"""
Construction du simulateur Excel IFC / IFC+ — SENIA S.A. (entreprise sénégalaise fictive)
Méthodologie : FICHE-IFC-TP (méthode prospective, convention CCNI Sénégal, table CIMA H, taux 3,5%)
Le classeur généré contient des formules Excel vivantes ; ce script calcule aussi
un miroir Python des résultats pour vérification.
"""
import datetime as dt
import json
import os
import random

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

# ----------------------------------------------------------------------------
# Hypothèses
# ----------------------------------------------------------------------------
DATE_EVAL = dt.date(2025, 12, 31)
AGE_RETRAITE = 60
TAUX_TECH = 0.035
TX_SAL = 0.02
TX_TURNOVER = 0.02
TX_LIC = 0.01
FRAIS_IFC = 0.02
FRAIS_IFCP = 0.02
DUREE_AMORT = 5
NB_EMP = 50

# ----------------------------------------------------------------------------
# Table CIMA H (Lx) — extraite de Evaluation/doc/Cours/CIMA-H__CIMA-F.xlsx
# ----------------------------------------------------------------------------
with open(os.path.join(REPO, "Evaluation/doc/Cours/CIMA-H__CIMA-F.xlsx"), "rb") as f:
    _wb = openpyxl.load_workbook(f, data_only=True)
_ws = _wb["CIMA-H "]
LX = {}
for row in _ws.iter_rows(min_row=3, values_only=True):
    if row[0] is not None and isinstance(row[0], (int, float)):
        LX[int(row[0])] = float(row[1])
AGE_MAX = max(LX)  # 106

# ----------------------------------------------------------------------------
# Barème CCNI (droits IFC, en % du salaire mensuel par année de service)
# 20 % / an les 5 premières années, 30 % de la 6e à la 10e, 40 % de la 11e à la
# 15e, 50 % de la 16e à la 20e, 60 % au-delà.
# ----------------------------------------------------------------------------
def droit_annuel(a):
    if a <= 0:
        return 0.0
    if a <= 5:
        return 0.20
    if a <= 10:
        return 0.30
    if a <= 15:
        return 0.40
    if a <= 20:
        return 0.50
    return 0.60

CUMUL = {0: 0.0}
for a in range(1, 61):
    CUMUL[a] = CUMUL[a - 1] + droit_annuel(a)

# ----------------------------------------------------------------------------
# Base de données : 50 salariés d'une entreprise sénégalaise (générée, seed fixe)
# ----------------------------------------------------------------------------
rng = random.Random(42)

NOMS = ["Diop", "Ndiaye", "Fall", "Sarr", "Ba", "Diallo", "Guèye", "Sy", "Sow",
        "Faye", "Mbaye", "Cissé", "Diouf", "Niang", "Kane", "Thiam", "Camara",
        "Seck", "Wade", "Mendy", "Sané", "Badji", "Ndoye", "Samb", "Dieng",
        "Lô", "Ka", "Touré", "Gning", "Diagne"]
PRENOMS_H = ["Mamadou", "Abdoulaye", "Ousmane", "Ibrahima", "Cheikh", "Moussa",
             "Alioune", "Serigne", "Pape", "Modou", "Assane", "Omar", "Souleymane",
             "Babacar", "Malick", "Amadou", "Idrissa", "Lamine", "Boubacar", "Saliou"]
PRENOMS_F = ["Fatou", "Aminata", "Awa", "Khady", "Ndèye", "Mariama", "Adja",
             "Coumba", "Astou", "Dieynaba", "Sokhna", "Rokhaya", "Aïssatou",
             "Bineta", "Yacine", "Maguette", "Salimata", "Ramatoulaye"]

CATEGORIES = (
    [("Cadre dirigeant", 1_800_000, 3_200_000)] * 3
    + [("Cadre", 750_000, 1_600_000)] * 9
    + [("Agent de maîtrise", 350_000, 700_000)] * 13
    + [("Employé", 150_000, 330_000)] * 15
    + [("Ouvrier", 75_000, 140_000)] * 10
)

def rand_date(y1, y2):
    y = rng.randint(y1, y2)
    m = rng.randint(1, 12)
    d = rng.randint(1, 28)
    return dt.date(y, m, d)

employes = []
for i, (cat, smin, smax) in enumerate(CATEGORIES, start=1):
    sexe = rng.choice(["M", "M", "F"])  # ~2/3 hommes
    prenom = rng.choice(PRENOMS_H if sexe == "M" else PRENOMS_F)
    nom = rng.choice(NOMS)
    # âge à l'évaluation entre 25 et 58 ans (cadres dirigeants plus âgés)
    age_min, age_max_ = (40, 58) if cat == "Cadre dirigeant" else (25, 58)
    age = rng.randint(age_min, age_max_)
    naissance = rand_date(DATE_EVAL.year - age - 1, DATE_EVAL.year - age)
    # embauche entre 21 ans et l'âge actuel (au plus tôt 1995, au plus tard mi-2025)
    emb_min = max(naissance.year + 21, 1995)
    emb_max = DATE_EVAL.year - 1
    embauche = rand_date(emb_min, max(emb_min, emb_max))
    if embauche <= naissance:
        embauche = dt.date(naissance.year + 21, embauche.month, embauche.day)
    sal_mensuel = round(rng.uniform(smin, smax) / 100) * 100
    employes.append(dict(mat=i, nom=nom, prenom=prenom, cat=cat, sexe=sexe,
                         naissance=naissance, embauche=embauche,
                         sal_mensuel=sal_mensuel))

# tri par date de naissance (comme la base KOSSAM)
employes.sort(key=lambda e: e["naissance"])
for i, e in enumerate(employes, start=1):
    e["mat"] = i

# ----------------------------------------------------------------------------
# Miroir Python des calculs (mêmes conventions que les formules Excel)
# ----------------------------------------------------------------------------
def yearfrac(d1, d2):
    return (d2 - d1).days / 365.25

def edate_years(d, years):
    return dt.date(d.year + years, d.month, d.day)

def cumul_interp(anc):
    k = int(anc)
    return CUMUL[k] + (anc - k) * droit_annuel(k + 1)

res_rows = []
for e in employes:
    SA = e["sal_mensuel"] * 12
    anc = yearfrac(e["embauche"], DATE_EVAL)
    age_exact = yearfrac(e["naissance"], DATE_EVAL)
    age = int(age_exact)
    ret = edate_years(e["naissance"], AGE_RETRAITE)
    n = max(yearfrac(DATE_EVAL, ret), 0.0)
    sal_final = SA / 12 * (1 + TX_SAL) ** max(AGE_RETRAITE - age, 0)
    anc_terme = anc + n
    droits = cumul_interp(anc_terme) * sal_final
    pvie = LX[AGE_RETRAITE] / LX[age] if age < AGE_RETRAITE else 1.0
    coef = 1 / (1 + TAUX_TECH) ** (n + 0.5)
    vp = droits * pvie * 1.0 * coef
    dette = vp * anc / anc_terme
    charge = vp / anc_terme
    ilic = 0.0 if n <= 0 else CUMUL[int(anc)] * SA / 12
    u = (1 - TX_TURNOVER) / (1 + TAUX_TECH)
    passif_il = 0.0 if n <= 0 else ilic * TX_LIC * (1 - u ** n) / (1 - u)
    base_if = SA / 24
    i_if = 0.0 if n <= 0 else CUMUL[int(anc)] * base_if
    cap_dc = ilic + i_if
    qx = (LX[age] - LX[age + 1]) / LX[age]
    prime_pure = cap_dc * qx
    res_rows.append(dict(SA=SA, anc=anc, age_exact=age_exact, age=age, ret=ret,
                         n=n, sal_final=sal_final, anc_terme=anc_terme,
                         droits=droits, pvie=pvie, coef=coef, vp=vp,
                         dette=dette, charge=charge, ilic=ilic,
                         passif_il=passif_il, base_if=base_if, i_if=i_if,
                         cap_dc=cap_dc, qx=qx, prime_pure=prime_pure))

tot = lambda k: sum(r[k] for r in res_rows)
masse = tot("SA")
engagement = tot("vp")
dette_act = tot("dette")
charge_norm = tot("charge")
passif_il = tot("passif_il")
cap_dc = tot("cap_dc")
prime_pure_ifcp = tot("prime_pure")
prime_comm_ifcp = prime_pure_ifcp / (1 - FRAIS_IFCP)
prime_unique = engagement + passif_il
amort = dette_act / DUREE_AMORT
prime_annuelle = amort + charge_norm

synth = {
    "effectif": NB_EMP,
    "masse_salariale": masse,
    "age_moyen": sum(r["age_exact"] for r in res_rows) / NB_EMP,
    "anciennete_moyenne": sum(r["anc"] for r in res_rows) / NB_EMP,
    "duree_residuelle_moyenne": sum(r["n"] for r in res_rows) / NB_EMP,
    "engagement_global": engagement,
    "dette_actuarielle": dette_act,
    "charge_normale": charge_norm,
    "passif_IL": passif_il,
    "prime_unique_opt1": prime_unique,
    "amortissement_dette_5ans": amort,
    "prime_annuelle_opt3": prime_annuelle,
    "taux_prime_uniforme_masse": prime_annuelle / masse,
    "capitaux_deces_IFCplus": cap_dc,
    "prime_pure_IFCplus": prime_pure_ifcp,
    "prime_commerciale_IFCplus": prime_comm_ifcp,
    "taux_prime_IFCplus_capitaux": prime_comm_ifcp / cap_dc,
    "taux_prime_IFCplus_masse": prime_comm_ifcp / masse,
}

# Sensibilité au taux technique et au taux d'évolution des salaires
def scenario(i_tech, ts):
    eng = det = ch = 0.0
    for e, r in zip(employes, res_rows):
        SA = e["sal_mensuel"] * 12
        sal_final = SA / 12 * (1 + ts) ** max(AGE_RETRAITE - r["age"], 0)
        droits = cumul_interp(r["anc_terme"]) * sal_final
        vp = droits * r["pvie"] / (1 + i_tech) ** (r["n"] + 0.5)
        eng += vp
        det += vp * r["anc"] / r["anc_terme"]
        ch += vp / r["anc_terme"]
    return eng, det, ch

sens = {}
for lbl, (it, ts) in {
    "central (i=3,5% ; s=2%)": (0.035, 0.02),
    "i=3,0%": (0.03, 0.02),
    "i=4,0%": (0.04, 0.02),
    "s=3%": (0.035, 0.03),
    "s=1%": (0.035, 0.01),
}.items():
    sens[lbl] = scenario(it, ts)

# ============================================================================
# Construction du classeur Excel
# ============================================================================
wb = openpyxl.Workbook()

# --- styles (sans couleur : gras + bordures noires uniquement) -------------
FMT_INT = "#,##0"
FMT_DEC2 = "#,##0.00"
FMT_PCT = "0.00%"
FMT_DATE = "DD/MM/YYYY"

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
H_FONT = Font(bold=True, size=10)
T_FONT = Font(bold=True, size=14)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

def style_header(ws, row, c1, c2):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = H_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def section(ws, row, text, c1=1, c2=3):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = Font(bold=True, size=11)
    cell.alignment = CENTER
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).border = BORDER

# --- Feuille Hypotheses -----------------------------------------------------
ws = wb.active
ws.title = "Hypotheses"
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
ws["A1"] = "SIMULATEUR IFC / IFC+ : ÉVALUATION ACTUARIELLE DU PASSIF SOCIAL"
ws["A1"].font = T_FONT
ws["A1"].alignment = CENTER

rows_h = [
    ("Entreprise", "SENIA S.A. (Sénégalaise des Industries Alimentaires)", None, None),
    ("Localisation", "Zone industrielle de Dakar, Sénégal", None, None),
    ("Convention collective", "Convention Collective Nationale Interprofessionnelle (CCNI) du Sénégal", None, None),
    ("Date d'évaluation", DATE_EVAL, "DateEval", FMT_DATE),
    ("Âge de départ à la retraite (IPRES)", AGE_RETRAITE, "AgeRetraite", "0"),
    ("Taux technique (actualisation)", TAUX_TECH, "TauxTech", FMT_PCT),
    ("Taux d'évolution des salaires", TX_SAL, "TxSal", FMT_PCT),
    ("Taux de turn-over", TX_TURNOVER, "TxTurnover", FMT_PCT),
    ("Taux annuel de licenciement (par défaut)", TX_LIC, "TxLic", FMT_PCT),
    ("Frais de gestion IFC", FRAIS_IFC, "FraisIFC", FMT_PCT),
    ("Frais de gestion IFC+", FRAIS_IFCP, "FraisIFCP", FMT_PCT),
    ("Durée d'amortissement de la dette actuarielle (ans)", DUREE_AMORT, "DureeAmort", "0"),
    ("Taux d'imposition des sociétés (info)", 0.30, None, FMT_PCT),
    ("SMIG mensuel légal (info, FCFA)", 43_440, None, FMT_INT),
    ("Table de mortalité", "CIMA H (art. 338 du Code CIMA)", None, None),
    ("Méthode d'évaluation", "Méthode prospective (droits projetés au terme, prorata d'ancienneté)", None, None),
]
r = 3
for label, val, name, fmt in rows_h:
    ws.cell(row=r, column=1, value=label).font = Font(bold=True)
    c = ws.cell(row=r, column=2, value=val)
    c.alignment = LEFT
    if fmt:
        c.number_format = fmt
    if name:
        wb.defined_names[name] = DefinedName(name, attr_text=f"Hypotheses!$B${r}")
    r += 1
ws.column_dimensions["A"].width = 48
ws.column_dimensions["B"].width = 62
for rr in range(3, r):
    ws.cell(row=rr, column=1).border = BORDER
    ws.cell(row=rr, column=2).border = BORDER

# --- Feuille BD --------------------------------------------------------------
ws = wb.create_sheet("BD")
headers = ["Matricule", "Nom et Prénoms", "Catégorie", "Sexe",
           "Date de naissance", "Date d'embauche",
           "Salaire mensuel brut (FCFA)", "Salaire annuel brut (FCFA)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header(ws, 1, 1, len(headers))
for i, e in enumerate(employes, start=2):
    ws.cell(row=i, column=1, value=e["mat"])
    ws.cell(row=i, column=2, value=f"{e['nom']} {e['prenom']}")
    ws.cell(row=i, column=3, value=e["cat"])
    ws.cell(row=i, column=4, value=e["sexe"])
    ws.cell(row=i, column=5, value=e["naissance"]).number_format = FMT_DATE
    ws.cell(row=i, column=6, value=e["embauche"]).number_format = FMT_DATE
    ws.cell(row=i, column=7, value=e["sal_mensuel"]).number_format = FMT_INT
    c = ws.cell(row=i, column=8, value=f"=G{i}*12")
    c.number_format = FMT_INT
    for j in range(1, 9):
        ws.cell(row=i, column=j).border = BORDER
for col, w in zip("ABCDEFGH", [10, 26, 18, 6, 15, 15, 16, 16]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- Feuille Baremes ---------------------------------------------------------
ws = wb.create_sheet("Baremes")
ws["A1"], ws["B1"], ws["C1"] = "Ancienneté (années)", "Droit annuel (% du salaire mensuel)", "Cumul (%)"
style_header(ws, 1, 1, 3)
ws["E1"] = ("Barème CCNI Sénégal : 20 %/an (années 1 à 5), 30 % (6 à 10), "
            "40 % (11 à 15), 50 % (16 à 20), 60 % au-delà.")
ws["E1"].font = Font(italic=True)
for a in range(0, 61):
    i = a + 2
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2,
            value=f"=IF(A{i}=0,0,IF(A{i}<=5,0.2,IF(A{i}<=10,0.3,IF(A{i}<=15,0.4,IF(A{i}<=20,0.5,0.6)))))")
    ws.cell(row=i, column=3, value=(0 if a == 0 else f"=C{i-1}+B{i}"))
    ws.cell(row=i, column=2).number_format = "0%"
    ws.cell(row=i, column=3).number_format = "0%"
    for j in range(1, 4):
        ws.cell(row=i, column=j).border = BORDER
for col, w in zip("ABC", [18, 30, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- Feuille Table_CIMA_H ----------------------------------------------------
ws = wb.create_sheet("Table_CIMA_H")
for j, h in enumerate(["Age x", "Lx", "dx", "qx", "Dx", "Nx", "Cx", "Mx"], start=1):
    ws.cell(row=1, column=j, value=h)
style_header(ws, 1, 1, 8)
last = AGE_MAX + 2  # dernière ligne
for a in range(0, AGE_MAX + 1):
    i = a + 2
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=LX[a])
    ws.cell(row=i, column=3, value=(f"=B{i}-B{i+1}" if a < AGE_MAX else f"=B{i}"))
    ws.cell(row=i, column=4, value=f"=IF(B{i}>0,C{i}/B{i},\"\")")
    ws.cell(row=i, column=5, value=f"=B{i}/(1+TauxTech)^A{i}")
    ws.cell(row=i, column=6, value=f"=SUM(E{i}:$E${last})")
    ws.cell(row=i, column=7, value=f"=C{i}/(1+TauxTech)^(A{i}+0.5)")
    ws.cell(row=i, column=8, value=f"=SUM(G{i}:$G${last})")
    ws.cell(row=i, column=2).number_format = FMT_INT
    ws.cell(row=i, column=3).number_format = FMT_INT
    ws.cell(row=i, column=4).number_format = "0.000000"
    for j in (5, 6, 7, 8):
        ws.cell(row=i, column=j).number_format = FMT_INT
for col, w in zip("ABCDEFGH", [8, 12, 10, 12, 12, 14, 10, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# --- Feuille Calculs ---------------------------------------------------------
ws = wb.create_sheet("Calculs")
headers = [
    "Matricule", "Date de naissance", "Date d'embauche", "Salaire annuel brut",
    "(1) Ancienneté à la date d'évaluation", "(2) Âge exact", "Âge (entier)",
    "Date de départ à la retraite", "Durée résiduelle n (années)",
    "(3) Salaire final mensuel", "(4) Ancienneté au terme", "(5) Droits IFC",
    "(6) Probabilité vie", "(7) Probabilité présence", "(9) Coefficient d'actualisation",
    "(10) VP des prestations futures", "(11) Dette actuarielle", "(12) Charge normale",
    "(13) Indemnité de licenciement", "(14) Passif IL", "(15) Demi-salaire mensuel (IF)",
    "(16) Indemnité IF", "(17) Capital décès total (IFC+)", "qx (âge actuel)",
    "Prime pure IFC+",
]
for j, h in enumerate(headers, start=1):
    ws.cell(row=1, column=j, value=h)
style_header(ws, 1, 1, len(headers))
ws.row_dimensions[1].height = 45

N1, N2 = 2, NB_EMP + 1
for i in range(N1, N2 + 1):
    F = {
        "A": f"=BD!A{i}",
        "B": f"=BD!E{i}",
        "C": f"=BD!F{i}",
        "D": f"=BD!H{i}",
        "E": f"=(DateEval-C{i})/365.25",
        "F": f"=(DateEval-B{i})/365.25",
        "G": f"=TRUNC(F{i})",
        "H": f"=EDATE(B{i},12*AgeRetraite)",
        "I": f"=MAX((H{i}-DateEval)/365.25,0)",
        "J": f"=D{i}/12*(1+TxSal)^MAX(AgeRetraite-G{i},0)",
        "K": f"=E{i}+I{i}",
        "L": f"=(VLOOKUP(TRUNC(K{i}),Baremes!$A:$C,3,TRUE)"
             f"+(K{i}-TRUNC(K{i}))*VLOOKUP(TRUNC(K{i})+1,Baremes!$A:$B,2,TRUE))*J{i}",
        "M": f"=IF(G{i}<AgeRetraite,VLOOKUP(AgeRetraite,Table_CIMA_H!$A:$B,2,TRUE)"
             f"/VLOOKUP(G{i},Table_CIMA_H!$A:$B,2,TRUE),1)",
        "N": "=1",
        "O": f"=1/(1+TauxTech)^(I{i}+0.5)",
        "P": f"=L{i}*M{i}*N{i}*O{i}",
        "Q": f"=P{i}*E{i}/K{i}",
        "R": f"=P{i}/K{i}",
        "S": f"=IF(I{i}<=0,0,VLOOKUP(TRUNC(E{i}),Baremes!$A:$C,3,TRUE)*D{i}/12)",
        "T": f"=IF(I{i}<=0,0,S{i}*TxLic*(1-((1-TxTurnover)/(1+TauxTech))^I{i})"
             f"/(1-(1-TxTurnover)/(1+TauxTech)))",
        "U": f"=D{i}/24",
        "V": f"=IF(I{i}<=0,0,VLOOKUP(TRUNC(E{i}),Baremes!$A:$C,3,TRUE)*U{i})",
        "W": f"=S{i}+V{i}",
        "X": f"=(VLOOKUP(G{i},Table_CIMA_H!$A:$B,2,TRUE)"
             f"-VLOOKUP(G{i}+1,Table_CIMA_H!$A:$B,2,TRUE))/VLOOKUP(G{i},Table_CIMA_H!$A:$B,2,TRUE)",
        "Y": f"=W{i}*X{i}",
    }
    for col, formula in F.items():
        cell = ws[f"{col}{i}"]
        cell.value = formula
        cell.border = BORDER
    ws[f"B{i}"].number_format = FMT_DATE
    ws[f"C{i}"].number_format = FMT_DATE
    ws[f"H{i}"].number_format = FMT_DATE
    for col in "DJLPQRSTUVWY":
        ws[f"{col}{i}"].number_format = FMT_INT
    for col in "EFIK":
        ws[f"{col}{i}"].number_format = FMT_DEC2
    ws[f"G{i}"].number_format = "0"
    for col in "MNO":
        ws[f"{col}{i}"].number_format = "0.0000"
    ws[f"X{i}"].number_format = "0.000000"
for j in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(j)].width = 15
ws.freeze_panes = "B2"

# ligne des totaux
tr = N2 + 1
ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
for col in "DPQRSTWY":
    c = ws[f"{col}{tr}"]
    c.value = f"=SUM({col}{N1}:{col}{N2})"
    c.number_format = FMT_INT
    c.font = Font(bold=True)
    c.border = BORDER

# --- Feuille Resultats -------------------------------------------------------
ws = wb.create_sheet("Resultats")
ws.sheet_view.showGridLines = False
ws.merge_cells("A1:C1")
ws["A1"] = "SYNTHÈSE DE L'ÉVALUATION AU 31/12/2025 : SENIA S.A."
ws["A1"].font = T_FONT
ws["A1"].alignment = CENTER

CA = f"Calculs!$D${N1}:$D${N2}"
def put(row, label, formula, fmt=FMT_INT, bold=False):
    ws.cell(row=row, column=1, value=label)
    c = ws.cell(row=row, column=2, value=formula)
    c.number_format = fmt
    if bold:
        ws.cell(row=row, column=1).font = Font(bold=True)
        c.font = Font(bold=True)
    for j in (1, 2):
        ws.cell(row=row, column=j).border = BORDER

r = 3
section(ws, r, "STATISTIQUES DU GROUPE"); r += 1
put(r, "Effectif", f"=COUNT(Calculs!A{N1}:A{N2})", "0"); r += 1
put(r, "Masse salariale annuelle (FCFA)", f"=SUM({CA})"); r += 1
put(r, "Âge moyen du groupe (années)", f"=AVERAGE(Calculs!F{N1}:F{N2})", FMT_DEC2); r += 1
put(r, "Ancienneté moyenne (années)", f"=AVERAGE(Calculs!E{N1}:E{N2})", FMT_DEC2); r += 1
put(r, "Durée résiduelle moyenne d'activité (années)", f"=AVERAGE(Calculs!I{N1}:I{N2})", FMT_DEC2); r += 2

section(ws, r, "RÉSULTATS IFC (FCFA)"); r += 1
r_eng = r
put(r, "Engagement global (VP des prestations futures)", f"=SUM(Calculs!P{N1}:P{N2})", bold=True); r += 1
r_dette = r
put(r, "Dette actuarielle", f"=SUM(Calculs!Q{N1}:Q{N2})", bold=True); r += 1
r_charge = r
put(r, "Charge normale de l'exercice", f"=SUM(Calculs!R{N1}:R{N2})", bold=True); r += 1
r_pil = r
put(r, "Passif de l'indemnité de licenciement", f"=SUM(Calculs!T{N1}:T{N2})", bold=True); r += 2

section(ws, r, "MODALITÉS DE FINANCEMENT"); r += 1
put(r, "Option 1 : prime unique (engagement global + passif IL)",
    f"=B{r_eng}+B{r_pil}", bold=True); r += 1
put(r, "Option 2 : prime initiale (dette actuarielle), puis charge normale annuelle",
    f"=B{r_dette}", bold=True); r += 1
put(r, "        dont charge normale de l'exercice suivant", f"=B{r_charge}"); r += 1
r_amort = r
put(r, "Option 3 : amortissement de la dette actuarielle (par an)",
    f"=B{r_dette}/DureeAmort"); r += 1
r_pa = r
put(r, "Option 3 : prime annuelle sur la durée d'amortissement",
    f"=B{r_amort}+B{r_charge}", bold=True); r += 1
put(r, "Taux de prime uniforme (en % de la masse salariale)",
    f"=B{r_pa}/SUM({CA})", FMT_PCT); r += 2

section(ws, r, "RÉSULTATS IFC+ (couverture décès, FCFA)"); r += 1
r_cap = r
put(r, "Capitaux garantis (indemnités décès CCNI)", f"=SUM(Calculs!W{N1}:W{N2})", bold=True); r += 1
r_pp = r
put(r, "Prime pure annuelle", f"=SUM(Calculs!Y{N1}:Y{N2})"); r += 1
r_pc = r
put(r, "Prime commerciale annuelle (chargée)", f"=B{r_pp}/(1-FraisIFCP)", bold=True); r += 1
put(r, "Taux de prime sur les capitaux garantis", f"=B{r_pc}/B{r_cap}", FMT_PCT); r += 1
put(r, "Taux de prime sur la masse salariale", f"=B{r_pc}/SUM({CA})", FMT_PCT); r += 1

ws.column_dimensions["A"].width = 58
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 8

# ordre des feuilles
wb.move_sheet("Resultats", offset=-(len(wb.sheetnames) - 1 - 1))  # après Hypotheses ? non
wb._sheets = [wb["Hypotheses"], wb["BD"], wb["Baremes"], wb["Table_CIMA_H"],
              wb["Calculs"], wb["Resultats"]]

wb.properties.creator = "Groupe 1 - ISE3 ENSAE"
wb.properties.title = "Simulateur IFC - SENIA S.A."
out = os.path.join(HERE, "simulateur_IFC_SENIA.xlsx")
wb.save(out)
print("Classeur écrit :", out)

# ----------------------------------------------------------------------------
# Sorties pour la note technique
# ----------------------------------------------------------------------------
print("\n================ SYNTHÈSE (miroir Python) ================")
for k, v in synth.items():
    if isinstance(v, float) and abs(v) > 1000:
        print(f"{k:40s} {v:>20,.0f}")
    else:
        print(f"{k:40s} {v:>20}" if not isinstance(v, float) else f"{k:40s} {v:>20.4f}")

print("\n================ SENSIBILITÉ ================")
for lbl, (eng, det, ch) in sens.items():
    print(f"{lbl:28s} engagement={eng:>15,.0f}  dette={det:>15,.0f}  charge={ch:>13,.0f}")

with open(os.path.join(HERE, "synthese.json"), "w") as f:
    json.dump({"synthese": synth,
               "sensibilite": {k: list(v) for k, v in sens.items()},
               "employes": [{**{kk: (str(vv) if isinstance(vv, dt.date) else vv)
                                for kk, vv in e.items()}} for e in employes]},
              f, ensure_ascii=False, indent=1, default=str)
print("\nsynthese.json écrit.")
